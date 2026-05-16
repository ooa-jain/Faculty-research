import os
import json
import io
from datetime import datetime, timezone, timedelta
from functools import wraps

from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, session, flash, send_file
)
from pymongo import MongoClient, DESCENDING
from pymongo.errors import ConnectionFailure, ServerSelectionTimeoutError
from bson import ObjectId
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "fallback-secret-key")

# ── Session & Security Config ─────────────────────────────────────────────────
app.config.update(
    PERMANENT_SESSION_LIFETIME=timedelta(hours=2),
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
)

# ── Rate Limiting ─────────────────────────────────────────────────────────────
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address
    limiter = Limiter(
        app=app,
        key_func=get_remote_address,
        default_limits=["500 per day", "100 per hour"],
        storage_uri="memory://",
    )
    LIMITER_AVAILABLE = True
except ImportError:
    LIMITER_AVAILABLE = False
    print("WARNING: flask-limiter not installed. Run: pip install Flask-Limiter")

# ── MongoDB with connection pooling ───────────────────────────────────────────
try:
    client = MongoClient(
        os.getenv("MONGO_URI"),
        maxPoolSize=50,
        minPoolSize=5,
        serverSelectionTimeoutMS=5000,
        connectTimeoutMS=10000,
        socketTimeoutMS=45000,
        retryWrites=True,
    )
    client.admin.command('ping')
    print("MongoDB connected successfully.")
except Exception as e:
    print(f"WARNING: MongoDB connection issue: {e}")

db = client.get_default_database()
responses_col = db["survey_responses"]
settings_col = db["settings"]

# ── Static file caching ───────────────────────────────────────────────────────
@app.after_request
def add_cache_headers(response):
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'public, max-age=3600'
    else:
        response.headers['Cache-Control'] = 'no-store'
    return response

# ── Helpers ───────────────────────────────────────────────────────────────────
class JSONEncoder(json.JSONEncoder):
    def default(self, o):
        if isinstance(o, ObjectId):
            return str(o)
        if isinstance(o, datetime):
            return o.isoformat()
        return super().default(o)

app.json_encoder = JSONEncoder


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


# ── Survey Routes ─────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("survey.html")


@app.route("/api/submit", methods=["POST"])
def submit_survey():
    payload = request.get_json(force=True)
    if not payload:
        return jsonify({"ok": False, "error": "Empty payload"}), 400

    consent = payload.get("consent", {})
    if not consent.get("agreed"):
        return jsonify({"ok": False, "error": "Consent not given"}), 400

    payload_str = json.dumps(payload)
    if len(payload_str) > 100_000:
        return jsonify({"ok": False, "error": "Payload too large"}), 413

    doc = {
        **payload,
        "submitted_at": datetime.now(timezone.utc),
        "ip": request.remote_addr,
    }

    try:
        result = responses_col.insert_one(doc)
        return jsonify({"ok": True, "id": str(result.inserted_id)}), 201
    except (ConnectionFailure, ServerSelectionTimeoutError) as e:
        app.logger.error(f"MongoDB connection error on submit: {e}")
        return jsonify({"ok": False, "error": "Database temporarily unavailable. Please retry."}), 503
    except Exception as e:
        app.logger.error(f"DB insert failed: {e}")
        return jsonify({"ok": False, "error": "Server error. Please retry."}), 500


# ── Health check ──────────────────────────────────────────────────────────────
@app.route("/health")
def health():
    try:
        client.admin.command('ping')
        return jsonify({"status": "ok", "db": "connected"}), 200
    except Exception as e:
        return jsonify({"status": "degraded", "db": str(e)}), 503


# ── Admin Routes ──────────────────────────────────────────────────────────────
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        if (username == os.getenv("ADMIN_USERNAME", "admin") and
                password == os.getenv("ADMIN_PASSWORD", "admin123")):
            session["admin_logged_in"] = True
            session.permanent = True
            return redirect(url_for("admin_dashboard"))
        flash("Invalid credentials", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    try:
        total = responses_col.count_documents({})
        consented = responses_col.count_documents({"consent.agreed": True})
        
        settings = settings_col.find_one({"_id": "global"}) or {"target_n": 80}
        target_n = settings.get("target_n", 80)

        pipeline_designation = [
            {"$group": {"_id": "$identity.designation", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}
        ]
        designation_data = list(responses_col.aggregate(pipeline_designation))

        pipeline_domain = [
            {"$group": {"_id": "$research.domain", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        domain_data = list(responses_col.aggregate(pipeline_domain))

        pipeline_ai = [
            {"$unwind": "$aitools.toolsUsed"},
            {"$group": {"_id": "$aitools.toolsUsed", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}},
            {"$limit": 10}
        ]
        ai_tools_data = list(responses_col.aggregate(pipeline_ai))

        pipeline_comfort = [
            {"$group": {"_id": "$aitools.overallComfort", "count": {"$sum": 1}}},
            {"$sort": {"_id": 1}}
        ]
        comfort_data = list(responses_col.aggregate(pipeline_comfort))

        recent = list(responses_col.find(
            {}, {"identity": 1, "submitted_at": 1, "research.domain": 1,
                 "aitools.overallComfort": 1, "consent": 1}
        ).sort("submitted_at", DESCENDING).limit(10))

        return render_template(
            "admin_dashboard.html",
            total=total, consented=consented,
            designation_data=designation_data,
            domain_data=domain_data,
            ai_tools_data=ai_tools_data,
            comfort_data=comfort_data,
            recent=recent,
            target_n=target_n,
        )
    except Exception as e:
        app.logger.error(f"Dashboard error: {e}")
        flash("Error loading dashboard data.", "error")
        return render_template("admin_dashboard.html",
            total=0, consented=0,
            designation_data=[], domain_data=[],
            ai_tools_data=[], comfort_data=[], recent=[], target_n=80)

@app.route("/admin/settings", methods=["POST"])
@admin_required
def admin_settings():
    data = request.json
    target_n = int(data.get("target_n", 80))
    settings_col.update_one(
        {"_id": "global"},
        {"$set": {"target_n": target_n}},
        upsert=True
    )
    return jsonify({"ok": True})


@app.route("/admin/responses")
@admin_required
def admin_responses():
    page = int(request.args.get("page", 1))
    per_page = 20
    skip = (page - 1) * per_page
    try:
        total = responses_col.count_documents({})
        docs = list(responses_col.find({}).sort(
            "submitted_at", DESCENDING).skip(skip).limit(per_page))
        pages = (total + per_page - 1) // per_page
    except Exception as e:
        app.logger.error(f"Responses list error: {e}")
        total, docs, pages = 0, [], 0
    return render_template(
        "admin_responses.html",
        docs=docs, page=page, pages=pages, total=total
    )


@app.route("/admin/response/<response_id>")
@admin_required
def admin_response_detail(response_id):
    try:
        doc = responses_col.find_one({"_id": ObjectId(response_id)})
    except Exception:
        flash("Invalid response ID.", "error")
        return redirect(url_for("admin_responses"))
    if not doc:
        flash("Response not found", "error")
        return redirect(url_for("admin_responses"))
    return render_template("admin_response_detail.html", doc=doc)


@app.route("/admin/export/json")
@admin_required
def export_json():
    try:
        docs = list(responses_col.find({}))
        for d in docs:
            d["_id"] = str(d["_id"])
            if "submitted_at" in d:
                d["submitted_at"] = d["submitted_at"].isoformat()
        response = app.response_class(
            json.dumps(docs, indent=2, ensure_ascii=False),
            mimetype="application/json",
            headers={"Content-Disposition": "attachment; filename=fdp_responses.json"}
        )
        return response
    except Exception as e:
        app.logger.error(f"JSON export error: {e}")
        flash("Export failed.", "error")
        return redirect(url_for("admin_dashboard"))


@app.route("/admin/export/excel")
@admin_required
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return "openpyxl not installed. Run: pip install openpyxl", 500

    try:
        docs = list(responses_col.find({}).sort("submitted_at", DESCENDING))
    except Exception as e:
        app.logger.error(f"Excel export DB error: {e}")
        flash("Export failed — database error.", "error")
        return redirect(url_for("admin_dashboard"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FDP Responses"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="00B4A6")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "#", "Submitted At", "Name", "Email", "Institution", "Department",
        "Designation", "Years Exp", "ORCID",
        "Research Domain", "Sub-Domain", "Keywords", "Paradigm", "Focus",
        "Pub Total", "Indexed", "h-Index", "Citations", "Patents",
        "AI Comfort", "AI Tools Used", "AI Concerns",
        "Innovation Project", "Innovation Stage",
        "Integration Level",
        "One Year Goal", "Commitment Action", "Additional Thoughts",
        "Consented"
    ]

    ws.row_dimensions[1].height = 30
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(14, len(h) + 2)

    def flat(val):
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        if val is None:
            return ""
        return str(val)

    for i, doc in enumerate(docs, 1):
        ident = doc.get("identity", {})
        res = doc.get("research", {})
        pub = doc.get("publication", {})
        ai = doc.get("aitools", {})
        inno = doc.get("innovation", {})
        integ = doc.get("integration", {})
        asp = doc.get("aspirations", {})
        con = doc.get("consent", {})
        submitted = doc.get("submitted_at", "")
        if hasattr(submitted, "strftime"):
            submitted = submitted.strftime("%d %b %Y %H:%M")

        row = [
            i, submitted,
            flat(ident.get("name")), flat(ident.get("email")),
            flat(ident.get("institution")), flat(ident.get("department")),
            flat(ident.get("designation")), flat(ident.get("yearsExp")),
            flat(ident.get("orcid")),
            flat(res.get("domain")), flat(res.get("subDomain")),
            flat(res.get("keywords")), flat(res.get("paradigm")),
            flat(res.get("focus")),
            flat(pub.get("total")), flat(pub.get("indexed")),
            flat(pub.get("hIndex")), flat(pub.get("citations")),
            flat(pub.get("patents")),
            flat(ai.get("overallComfort")), flat(ai.get("toolsUsed")),
            flat(ai.get("concerns")),
            flat(inno.get("hasProject")), flat(inno.get("stage")),
            flat(integ.get("integrationLevel")),
            flat(asp.get("oneYearGoal")), flat(asp.get("commitmentAction")),
            flat(asp.get("additionalThoughts")),
            "Yes" if con.get("agreed") else "No",
        ]
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i + 1].height = 18

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="fdp_responses.xlsx"
    )


@app.route("/admin/delete/<response_id>", methods=["POST"])
@admin_required
def admin_delete(response_id):
    try:
        responses_col.delete_one({"_id": ObjectId(response_id)})
        flash("Response deleted.", "success")
    except Exception as e:
        app.logger.error(f"Delete error: {e}")
        flash("Delete failed.", "error")
    return redirect(url_for("admin_responses"))


# ── Error handlers ────────────────────────────────────────────────────────────
@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not found"}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Internal server error"}), 500

@app.errorhandler(429)
def rate_limit_exceeded(e):
    return jsonify({"ok": False, "error": "Too many requests. Please wait and retry."}), 429


if __name__ == "__main__":
    app.run(debug=False, port=5000, threaded=True)
